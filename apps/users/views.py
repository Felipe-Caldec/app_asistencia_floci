"""
apps/users/views.py
"""

import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q

from .models import ProfesorProfile
from .forms import CrearProfesorForm, EditarProfesorForm


def es_admin(user):
    return user.is_superuser


def _filtrar_profesores(params):
    """Aplica filtros al queryset de profesores."""
    qs = User.objects.filter(is_superuser=False).select_related('profesor_profile').order_by('last_name', 'first_name')

    nombre      = params.get('nombre', '').strip()
    rut         = params.get('rut', '').strip()
    departamento = params.get('departamento', '').strip()

    if nombre:
        qs = qs.filter(Q(first_name__icontains=nombre) | Q(last_name__icontains=nombre))
    if rut:
        qs = qs.filter(username__icontains=rut)
    if departamento:
        qs = qs.filter(profesor_profile__departamento__icontains=departamento)

    return qs


@login_required
@user_passes_test(es_admin, login_url='/')
def lista_profesores(request):
    profesores = _filtrar_profesores(request.GET)
    return render(request, 'users/lista_profesores.html', {
        'profesores': profesores,
        'filtros': request.GET,
    })


@login_required
@user_passes_test(es_admin, login_url='/')
def exportar_profesores_excel(request):
    """Descarga la lista de profesores en Excel respetando los filtros activos."""
    profesores = _filtrar_profesores(request.GET)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Profesores'

    header_fill = PatternFill('solid', fgColor='1a2d7a')
    header_font = Font(color='FFFFFF', bold=True)
    headers = ['Usuario', 'Nombre', 'Apellido', 'Email', 'Departamento', 'Teléfono', 'Estado']

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for profesor in profesores:
        perfil = getattr(profesor, 'profesor_profile', None)
        ws.append([
            profesor.username,
            profesor.first_name,
            profesor.last_name,
            profesor.email,
            perfil.departamento if perfil else '',
            perfil.telefono if perfil else '',
            'Activo' if profesor.is_active else 'Inactivo',
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="profesores.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin, login_url='/')
def crear_profesor(request):
    if request.method == 'POST':
        form = CrearProfesorForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password1'])
            user.save()
            ProfesorProfile.objects.create(
                user=user,
                telefono=form.cleaned_data.get('telefono', ''),
                departamento=form.cleaned_data.get('departamento', '')
            )
            messages.success(request, f'Profesor {user.get_full_name()} creado exitosamente.')
            return redirect('lista_profesores')
    else:
        form = CrearProfesorForm()
    return render(request, 'users/crear_profesor.html', {'form': form})


@login_required
@user_passes_test(es_admin, login_url='/')
def editar_profesor(request, pk):
    profesor = get_object_or_404(User, pk=pk, is_superuser=False)
    perfil, _ = ProfesorProfile.objects.get_or_create(user=profesor)
    if request.method == 'POST':
        form = EditarProfesorForm(request.POST, instance=profesor)
        if form.is_valid():
            form.save()
            perfil.telefono = form.cleaned_data.get('telefono', '')
            perfil.departamento = form.cleaned_data.get('departamento', '')
            perfil.save()
            messages.success(request, 'Datos actualizados correctamente.')
            return redirect('lista_profesores')
    else:
        initial = {'telefono': perfil.telefono, 'departamento': perfil.departamento}
        form = EditarProfesorForm(instance=profesor, initial=initial)
    return render(request, 'users/editar_profesor.html', {'form': form, 'profesor': profesor})


@login_required
@user_passes_test(es_admin, login_url='/')
def toggle_profesor(request, pk):
    profesor = get_object_or_404(User, pk=pk, is_superuser=False)
    if request.method == 'POST':
        profesor.is_active = not profesor.is_active
        profesor.save()
        estado = 'activado' if profesor.is_active else 'desactivado'
        messages.success(request, f'Usuario {profesor.username} {estado}.')
    return redirect('lista_profesores')


@login_required
@user_passes_test(es_admin, login_url='/')
def carga_csv_profesores(request):
    """
    Carga masiva de profesores desde CSV.
    Columnas esperadas: username, first_name, last_name, email, password, departamento, telefono
    """
    if request.method == 'POST' and request.FILES.get('archivo_csv'):
        archivo = request.FILES['archivo_csv']
        try:
            contenido = archivo.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(contenido))
            reader.fieldnames = [f.strip().lower() for f in reader.fieldnames]

            creados = 0
            omitidos = 0
            errores = []

            for i, fila in enumerate(reader, start=2):
                username   = fila.get('username', '').strip()
                first_name = fila.get('first_name', '').strip()
                last_name  = fila.get('last_name', '').strip()
                email      = fila.get('email', '').strip()
                password   = fila.get('password', '').strip() or 'Cambiar123!'
                departamento = fila.get('departamento', '').strip()
                telefono   = fila.get('telefono', '').strip()

                if not username or not first_name or not last_name:
                    errores.append(f'Fila {i}: faltan campos obligatorios (username, first_name, last_name).')
                    continue

                if User.objects.filter(username=username).exists():
                    omitidos += 1
                    continue

                user = User.objects.create_user(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    password=password,
                )
                ProfesorProfile.objects.create(
                    user=user,
                    departamento=departamento,
                    telefono=telefono
                )
                creados += 1

            if creados:
                messages.success(request, f'{creados} profesor(es) importado(s). Contraseña por defecto: "Cambiar123!" (si no se especificó).')
            if omitidos:
                messages.warning(request, f'{omitidos} profesor(es) omitido(s) por username duplicado.')
            for e in errores[:5]:
                messages.error(request, e)

        except Exception as ex:
            messages.error(request, f'Error al procesar el archivo: {ex}')

        return redirect('lista_profesores')

    return redirect('crear_profesor')
