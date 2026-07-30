"""
apps/courses/views.py
"""

import csv
import io
import json
import openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill, Alignment

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q

from .models import Curso, Horario, Alumno, AsignacionAlumno
from .forms import CursoForm, HorarioForm, AlumnoForm, AsignacionAlumnoForm


def es_admin(user):
    return user.is_superuser


# Mapeo día Python (0=lunes) → código del modelo
DIA_MAP = {0: 'LUN', 1: 'MAR', 2: 'MIE', 3: 'JUE', 4: 'VIE', 5: 'SAB', 6: 'DOM'}


# ── DASHBOARD ────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    dia_codigo = DIA_MAP.get(date.today().weekday())
    dia_nombre = dict(Horario.DIAS_SEMANA).get(dia_codigo, dia_codigo)

    if request.user.is_superuser:
        # Admin ve TODOS los cursos con TODOS sus horarios, sin filtro de día
        cursos_qs = Curso.objects.filter(activo=True).distinct()

        cursos = []
        for curso in cursos_qs:
            horarios = curso.horarios.filter(activo=True).select_related('profesor')
            cursos.append({'curso': curso, 'horarios_hoy': horarios})

    else:
        # Profesor solo ve sus cursos con horario del día actual
        cursos_qs = Curso.objects.filter(
            activo=True,
            horarios__activo=True,
            horarios__dia_semana=dia_codigo,
            horarios__profesor=request.user
        ).distinct()

        cursos = []
        for curso in cursos_qs:
            horarios_hoy = curso.horarios.filter(
                activo=True, dia_semana=dia_codigo, profesor=request.user
            )
            cursos.append({'curso': curso, 'horarios_hoy': horarios_hoy})

    return render(request, 'courses/dashboard.html', {
        'cursos': cursos,
        'dia_nombre': dia_nombre,
        'dia_hoy': dia_codigo,
        'es_admin': request.user.is_superuser,
    })


# ── DETALLE CURSO ─────────────────────────────────────────────────────────────

@login_required
def detalle_curso(request, pk):
    curso = get_object_or_404(Curso, pk=pk, activo=True)
    if request.user.is_superuser:
        horarios = curso.horarios.filter(activo=True).select_related('profesor')
    else:
        dia_codigo = DIA_MAP.get(date.today().weekday())
        horarios = curso.horarios.filter(
            activo=True, profesor=request.user,
            dia_semana=dia_codigo
        ).select_related('profesor')
        if not horarios.exists():
            messages.error(request, 'No tienes acceso a este curso.')
            return redirect('dashboard')
    return render(request, 'courses/detalle_curso.html', {
        'curso': curso, 'horarios': horarios
    })


# ── API: horarios por curso (para el select dinámico de asignación) ──────────

@login_required
@user_passes_test(es_admin, login_url='/')
def horarios_por_curso(request, curso_pk):
    """
    Endpoint JSON que devuelve los horarios activos de un curso.
    Usado por el JS del formulario de asignación de alumnos.
    """
    horarios = Horario.objects.filter(
        curso_id=curso_pk, activo=True
    ).select_related('profesor').order_by('dia_semana', 'hora_inicio')

    data = [
        {
            'id': h.pk,
            'label': f"{h.get_dia_semana_display()} {h.hora_inicio.strftime('%H:%M')}–{h.hora_fin.strftime('%H:%M')} — {h.profesor_nombre()}"
        }
        for h in horarios
    ]
    return JsonResponse({'horarios': data})


# ── CURSOS CRUD ───────────────────────────────────────────────────────────────

@login_required
@user_passes_test(es_admin, login_url='/')
def crear_curso(request):
    if request.method == 'POST':
        form = CursoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Curso creado exitosamente.')
            return redirect('dashboard')
    else:
        form = CursoForm()
    return render(request, 'courses/form_curso.html', {'form': form, 'titulo': 'Crear Curso'})


@login_required
@user_passes_test(es_admin, login_url='/')
def editar_curso(request, pk):
    curso = get_object_or_404(Curso, pk=pk)
    if request.method == 'POST':
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Curso actualizado.')
            return redirect('dashboard')
    else:
        form = CursoForm(instance=curso)
    return render(request, 'courses/form_curso.html', {
        'form': form, 'titulo': 'Editar Curso', 'curso': curso
    })


# ── HORARIOS CRUD ─────────────────────────────────────────────────────────────

@login_required
@user_passes_test(es_admin, login_url='/')
def crear_horario(request):
    if request.method == 'POST':
        form = HorarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horario creado exitosamente.')
            return redirect('dashboard')
    else:
        form = HorarioForm()
    return render(request, 'courses/form_horario.html', {
        'form': form, 'titulo': 'Crear Horario'
    })


@login_required
@user_passes_test(es_admin, login_url='/')
def editar_horario(request, pk):
    horario = get_object_or_404(Horario, pk=pk)
    if request.method == 'POST':
        form = HorarioForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horario actualizado.')
            return redirect('detalle_curso', pk=horario.curso.pk)
    else:
        form = HorarioForm(instance=horario)
    return render(request, 'courses/form_horario.html', {
        'form': form, 'titulo': 'Editar Horario', 'horario': horario
    })


# ── ALUMNOS ───────────────────────────────────────────────────────────────────

def _filtrar_alumnos(params):
    qs = Alumno.objects.filter(activo=True).prefetch_related('asignaciones__curso')
    nombre = params.get('nombre', '').strip()
    rut    = params.get('rut', '').strip()
    genero = params.get('genero', '').strip()
    curso  = params.get('curso', '').strip()
    if nombre:
        qs = qs.filter(Q(nombre__icontains=nombre) | Q(apellido__icontains=nombre))
    if rut:
        qs = qs.filter(rut__icontains=rut)
    if genero:
        qs = qs.filter(genero=genero)
    if curso:
        qs = qs.filter(asignaciones__curso__id=curso)
    return qs.distinct()


@login_required
@user_passes_test(es_admin, login_url='/')
def lista_alumnos(request):
    alumnos = _filtrar_alumnos(request.GET)
    cursos  = Curso.objects.filter(activo=True).order_by('nombre')
    return render(request, 'courses/lista_alumnos.html', {
        'alumnos': alumnos,
        'cursos': cursos,
        'filtros': request.GET,
        'genero_choices': Alumno.GENERO_CHOICES,
    })


@login_required
@user_passes_test(es_admin, login_url='/')
def exportar_alumnos_excel(request):
    alumnos = _filtrar_alumnos(request.GET)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Alumnos'
    header_fill = PatternFill('solid', fgColor='1a2d7a')
    header_font = Font(color='FFFFFF', bold=True)
    headers = ['RUT', 'Apellido', 'Nombre', 'Género', 'Correo', 'Teléfono', 'Cursos']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for alumno in alumnos:
        cursos_str = ', '.join(a.curso.nombre for a in alumno.asignaciones.all())
        ws.append([
            alumno.rut, alumno.apellido, alumno.nombre,
            alumno.get_genero_display() if alumno.genero else '',
            alumno.correo, alumno.telefono, cursos_str,
        ])
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="alumnos.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(es_admin, login_url='/')
def crear_alumno(request):
    if request.method == 'POST':
        form = AlumnoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Alumno creado exitosamente.')
            return redirect('lista_alumnos')
    else:
        form = AlumnoForm()
    return render(request, 'courses/form_alumno.html', {
        'form': form, 'titulo': 'Crear Alumno'
    })


@login_required
@user_passes_test(es_admin, login_url='/')
def editar_alumno(request, pk):
    alumno = get_object_or_404(Alumno, pk=pk)
    if request.method == 'POST':
        form = AlumnoForm(request.POST, instance=alumno)
        if form.is_valid():
            form.save()
            messages.success(request, 'Alumno actualizado.')
            return redirect('lista_alumnos')
    else:
        form = AlumnoForm(instance=alumno)
    return render(request, 'courses/form_alumno.html', {
        'form': form, 'titulo': 'Editar Alumno', 'alumno': alumno
    })


@login_required
@user_passes_test(es_admin, login_url='/')
def carga_csv_alumnos(request):
    if request.method == 'POST' and request.FILES.get('archivo_csv'):
        archivo = request.FILES['archivo_csv']
        try:
            contenido = archivo.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(contenido))
            reader.fieldnames = [f.strip().lower() for f in reader.fieldnames]
            creados = omitidos = 0
            errores = []
            for i, fila in enumerate(reader, start=2):
                nombre   = fila.get('nombre', '').strip()
                apellido = fila.get('apellido', '').strip()
                rut      = fila.get('rut', '').strip()
                if not nombre or not apellido or not rut:
                    errores.append(f'Fila {i}: faltan campos obligatorios.')
                    continue
                genero   = fila.get('genero', '').strip().upper()
                correo   = fila.get('correo', '').strip()
                telefono = fila.get('telefono', '').strip()
                if genero not in ('M', 'F', 'O'):
                    genero = ''
                _, created = Alumno.objects.get_or_create(
                    rut=rut,
                    defaults={
                        'nombre': nombre, 'apellido': apellido,
                        'genero': genero, 'correo': correo, 'telefono': telefono,
                    }
                )
                if created:
                    creados += 1
                else:
                    omitidos += 1
            if creados:
                messages.success(request, f'{creados} alumno(s) importado(s).')
            if omitidos:
                messages.warning(request, f'{omitidos} alumno(s) omitido(s) por RUT duplicado.')
            for e in errores[:5]:
                messages.error(request, e)
        except Exception as ex:
            messages.error(request, f'Error al procesar el archivo: {ex}')
        return redirect('lista_alumnos')
    return redirect('crear_alumno')


# ── ASIGNACIONES ──────────────────────────────────────────────────────────────

@login_required
@user_passes_test(es_admin, login_url='/')
def asignar_alumno(request):
    if request.method == 'POST':
        form = AsignacionAlumnoForm(request.POST)
        if form.is_valid():
            alumno  = form.cleaned_data['alumno']
            curso   = form.cleaned_data['curso']
            horario = form.cleaned_data['horario']
            obj, created = AsignacionAlumno.objects.get_or_create(
                alumno=alumno, horario=horario,
                defaults={'curso': curso}
            )
            if created:
                messages.success(request, f'{alumno} asignado a {horario} correctamente.')
            else:
                messages.warning(request, f'{alumno} ya estaba asignado a ese horario.')
            return redirect('dashboard')
    else:
        form = AsignacionAlumnoForm()

    # Pasar todos los cursos al template para alimentar el JS
    cursos = Curso.objects.filter(activo=True).order_by('nombre')
    return render(request, 'courses/form_asignacion.html', {
        'form': form,
        'cursos': cursos,
    })


@login_required
@user_passes_test(es_admin, login_url='/')
def asignar_profesor_horario(request):
    horarios = Horario.objects.filter(activo=True).select_related('curso', 'profesor')
    return render(request, 'courses/asignar_profesor.html', {'horarios': horarios})
